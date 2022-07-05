import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value

    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] += 1
    else:
        # print("create new supplier")
        product_per_supplier[supplier_name] = 1

    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += (inventory * price)
    else:
        # print("create new supplier value")
        total_value_per_supplier[supplier_name] = inventory * price

    if inventory < 10:
        product_under_10_inv[int(product_num)] = int(inventory)

    product_list.cell(product_row, 5).value = inventory * price

print(product_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)
inv_file.save("inventory.xlsx")
